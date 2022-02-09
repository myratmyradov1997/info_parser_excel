from bs4 import BeautifulSoup
import requests
import openpyxl

fake = ''
fake_first = ''
counter = 1
mas = []
fake_name = ''
names_main = ''
options_main = ''
fake_options = ''

book = openpyxl.Workbook()
sheet = book.active


def search(names, options, fake_name,counter):
    if 'Поверхность' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=1).value = names
        sheet.cell(row=counter, column=2).value = options

    if 'Толщина(мм)' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=3).value = names
        sheet.cell(row=counter, column=4).value = options

    if 'Обработкакрая' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=5).value = names
        sheet.cell(row=counter, column=6).value = options

    if 'Основнойцвет' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=7).value = names
        sheet.cell(row=counter, column=8).value = options

    if 'Рисунок' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=9).value = names
        sheet.cell(row=counter, column=10).value = options

    if 'Стильколлекции' not in fake_name:
        pass
    else:
        sheet.cell(row=counter, column=11).value = names
        sheet.cell(row=counter, column=12).value = options
    pass


def send_data(names_main, options_main,counter):
    for i in range(3, 14):
        fake_name = str(names_main[i].text.replace(' ', '')).splitlines()
        names = fake_name[1]
        fake_options = str(options_main[i - 3].text.replace(' ', '')).splitlines()
        options = fake_options[1]
        search(names,options,fake_name,counter)

    pass


with open('id.txt', 'r') as f:
    for line in f:
        print(counter)
        # mas=[]
        try:
            a = line.rstrip('\n')
            url = "https://shop.italonceramica.ru/search/index.php?q=" + a
            request = requests.get(url)
            soup = BeautifulSoup(request.text, "html.parser")

            div_tovar = soup.find_all("div", class_="main-desc")
            link = div_tovar[0].find_all("a")
            fake_first = ("https://shop.italonceramica.ru" + link[0]["href"])

            url = fake_first
            request = requests.get(url)
            soup = BeautifulSoup(request.text, "html.parser")

            names_main = soup.find_all("div", class_="name")
            options_main = soup.find_all("div", class_="options")
            send_data(names_main, options_main, counter)
            counter += 1

        except IndexError:
            counter += 1

book.save('info.xlsx')
book.close()
